from pathlib import Path
from selenium import webdriver
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import time
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import StaleElementReferenceException
from selenium.common.exceptions import NoSuchElementException
import re
import openpyxl
from pathlib import Path
from openpyxl.utils import get_column_letter

#usual with selenium
url = 'https://www.imdb.com/search/title/?title_type=feature&release_date=2022-01-12,2022-02-12' #release_date time period can be anything
driver = webdriver.Firefox()
driver.maximize_window()
driver.get(url)
wait = WebDriverWait(driver, 10)
movieDict = {}
genreSet = set()

#creating excel sheet
wb = openpyxl.load_workbook(Path.cwd() / 'imdbScrapper.xlsx')
sheet = wb['Sheet']

#clearing sheet
for row in sheet:
    for cell in row:
        cell.value = None

#hits the "50 more" button until there is no more "50 more" buttons
while True:
  try:
    element = wait.until(EC.element_to_be_clickable((By.CLASS_NAME, 'ipc-btn.ipc-btn--single-padding.ipc-btn--center-align-content.ipc-btn--default-height.ipc-btn--core-base.ipc-btn--theme-base.ipc-btn--on-accent2.ipc-text-button.ipc-see-more__button')))
    driver.execute_script("arguments[0].click();", element)

  except (TimeoutException, StaleElementReferenceException):
    break

#soup = wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[2]/main/div[2]/div[4]/section/section/div/section/section/div[2]/div/section/div[2]/div[2]/ul/li[1]/div/div/div/div[1]/div[3]/button/svg"))).click()  
soup = driver.find_elements(By.CLASS_NAME, "ipc-icon-button.dli-info-icon.ipc-icon-button--base.ipc-icon-button--onAccent2")
x = open(Path.cwd() / 'x.txt', 'w')

for i in soup:
  driver.execute_script("arguments[0].click();", i)
  time.sleep(.3)
  while True:
    try:
      rating = driver.find_element(By.CLASS_NAME, "ipc-rating-star.ipc-rating-star--baseAlt.ipc-rating-star--imdb.btp_rt_ds")
      length = driver.find_element(By.XPATH, "/html/body/div[4]/div[2]/div/div[2]/div/div/div[1]/div[2]/ul[1]/li[2]")
      genres = driver.find_element(By.XPATH, "/html/body/div[4]/div[2]/div/div[2]/div/div/div[1]/div[2]/ul[2]")
      genreList = re.findall(r'(?:Sci-Fi|[A-Z][^A-Z]*)', str(genres.text))
      lengthInMinutes = re.findall(r'(\d+)', length.text)
      if len(lengthInMinutes) == 1:
        lengthInMinutes.insert(0, 0)
      lengthInMinutes = sum([int(minutes) * 60 if ind == 0 else int(minutes) for ind, minutes in enumerate(lengthInMinutes)])
      for genre in genreList:
        if genre not in movieDict:
          movieDict[genre] = [[], [], 0]
      
        movieDict[genre][0].append(lengthInMinutes)
        movieDict[genre][1].append(float(rating.text[0:3]))
        movieDict[genre][2] += 1
      wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "ipc-promptable-base__close"))).click()
      break
    except NoSuchElementException:
      wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "ipc-promptable-base__close"))).click()
      break

#gets averages
for k in movieDict.keys():
  for i in range(2):
    movieDict[k][i] = round(sum(movieDict[k][i]) / len(movieDict[k][i]), 2)

#excel script
sheet.cell(row = 1, column = 2).value = 'average movie duration'
sheet.cell(row = 1, column = 3).value = 'average rating'
sheet.cell(row = 1, column = 4).value = 'amount of movies'

for key_index, key in enumerate(movieDict.keys()):
  sheet.cell(row = key_index + 2, column = 1).value = key

for index, (k, v) in enumerate(movieDict.items()):
  for i in range(2, 5):
    sheet.cell(row = index + 2, column = i).value = v[i-2]

x.write(str(movieDict))
x.close()

wb.save('imdbScrapper.xlsx')
driver.close()





  